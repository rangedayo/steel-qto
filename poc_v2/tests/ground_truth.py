"""도면 정답지(xlsx) 로더 — 회귀 테스트용 결정론적 정답 데이터.

라운드 7 방침: 시트가 `도면N-기둥`·`도면N-보` 로 분리됐다.
시트 안 구조:
    1행 = 헤더: `도면명 | 분석 대상 | <부호1> | <부호2> | ... | 합계`
    2행~끝-1 = 도면 내 세부 도면별 부호 개수
    끝 행 = `합계` 행(SUM 수식, 신뢰 금지)

셀 값 해석 규칙:
    빈 셀(None, "")    → 해당 도면에 그 부호 없음 (0으로 간주)
    숫자가 아닌 값      → 0으로 간주
    `합계` 행/`합계` 열 → 집계용이므로 제외 (데이터 행을 직접 합산한다)
    `분석 대상` 열      → 메모용, 카운트와 무관

LLM 호출 등 비결정 요소 없이 순수하게 xlsx만 파싱한다.
"""
from __future__ import annotations

import os

import openpyxl

# tests/ground_truth.py → poc_v2/tests → poc_v2 → 프로젝트 루트
_HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(_HERE))
ANSWER_KEY_PATH = os.path.join(
    PROJECT_ROOT, "reference_materials", "도면_정답지.xlsx"
)
SYMBOL_RULES_PATH = os.path.join(
    PROJECT_ROOT, "config", "symbol_rules.yaml"
)

_TOTAL_LABEL = "합계"
_NON_SYMBOL_HEADERS = frozenset({"도면명", "분석 대상", _TOTAL_LABEL})
_CATEGORY_SUFFIXES = ("-기둥", "-보")


def load_text_height_filter(
    path: str | None = None,
) -> dict[str, float | None]:
    """symbol_rules.yaml 의 text_height_filter → {도면명: min_height}.

    min_height 가 null(YAML) 이면 None 으로, 숫자면 그대로 돌려준다.
    설정 파일이나 pyyaml 이 없으면 빈 dict (= 모든 도면 필터 미적용).
    """
    config = _load_config(path)
    raw = config.get("text_height_filter", {}) or {}
    return {
        drawing: (spec or {}).get("min_height")
        for drawing, spec in raw.items()
    }


def load_auto_policy_params(
    path: str | None = None,
) -> dict[str, float]:
    """symbol_rules.yaml 의 auto_policy_params → 자동 정책 판단 파라미터.

    라운드 6: 신호 3(규격 패턴) 임계값. 누락 시 auto_detect_policy 기본값.
    """
    config = _load_config(path)
    raw = config.get("auto_policy_params", {}) or {}
    return {
        "spec_pattern_threshold": float(raw.get("spec_pattern_threshold", 0.3)),
    }


def load_policy_override(
    drawing_name: str,
    path: str | None = None,
) -> dict[str, bool] | None:
    """symbol_rules.yaml 의 policy_override[drawing_name] 로드.

    라운드 6: 자동 판단을 강제로 덮어쓰는 비상용 설정. yaml 값이 null 이거나
    키가 없으면 None 을 돌려준다(= 자동 판단 사용). dict 면 두 플래그를 채운
    정규화된 dict 를 돌려준다.
    """
    config = _load_config(path)
    raw = (config.get("policy_override", {}) or {}).get(drawing_name)
    if not raw:
        return None
    return {
        "exclude_table_regions": bool(raw.get("exclude_table_regions", False)),
        "exclude_with_spec": bool(raw.get("exclude_with_spec", False)),
    }


def load_policy_p(
    path: str | None = None,
) -> dict[str, dict[str, bool]]:
    """[DEPRECATED — 라운드 6] symbol_rules.yaml 의 policy_p → {도면명: {플래그}}.

    라운드 5 정책 P 도면별 수동 분기 로더. 라운드 6 에서 policy_p 키는
    주석 처리되고 신호 2·3 은 auto_policy.auto_detect_policy 가 자동 판단한다.
    yaml 키가 폐기됐으므로 이 함수는 빈 dict 를 돌려준다. 호출 금지.
    """
    config = _load_config(path)
    raw = config.get("policy_p", {}) or {}
    result: dict[str, dict[str, bool]] = {}
    for drawing, spec in raw.items():
        spec = spec or {}
        result[drawing] = {
            "exclude_table_regions": bool(spec.get("exclude_table_regions", False)),
            "exclude_with_spec": bool(spec.get("exclude_with_spec", False)),
        }
    return result


def load_table_region_params(
    path: str | None = None,
) -> dict[str, float | int]:
    """symbol_rules.yaml 의 table_region_detection → 일람표 검출 파라미터.

    누락 시 detect_table_regions 기본값과 같은 값으로 채운다.
    """
    config = _load_config(path)
    raw = config.get("table_region_detection", {}) or {}
    return {
        "region_size_ratio": float(raw.get("region_size_ratio", 1 / 30)),
        "min_distinct_symbols": int(raw.get("min_distinct_symbols", 4)),
        "max_count_per_symbol": int(raw.get("max_count_per_symbol", 2)),
    }


def _load_config(path: str | None) -> dict:
    """symbol_rules.yaml 파싱. pyyaml·파일 부재 시 빈 dict."""
    try:
        import yaml  # noqa: PLC0415
    except ImportError:
        return {}
    cfg_path = path or SYMBOL_RULES_PATH
    if not os.path.exists(cfg_path):
        return {}
    with open(cfg_path, encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def _parse_cell_int(value: object) -> int:
    """셀 값을 정수 카운트로 해석. 빈 셀·비숫자는 0."""
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _drawing_name_from_sheet(sheet_name: str) -> tuple[str, str | None]:
    """`도면1-기둥` → (`도면1`, `기둥`). 접미사 없으면 (sheet_name, None)."""
    for suffix in _CATEGORY_SUFFIXES:
        if sheet_name.endswith(suffix):
            return sheet_name[: -len(suffix)], suffix[1:]
    return sheet_name, None


def _parse_sheet(worksheet) -> dict[str, int]:
    """시트 하나에서 부호별 합계를 데이터 행 직접 합산으로 산출."""
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = rows[0]
    symbol_columns: list[tuple[int, str]] = []
    for col_idx, raw_name in enumerate(header):
        if raw_name is None:
            continue
        label = str(raw_name).strip()
        if not label or label in _NON_SYMBOL_HEADERS:
            continue
        symbol_columns.append((col_idx, label))

    aggregated: dict[str, int] = {}
    for row in rows[1:]:
        if not row:
            continue
        raw_first = row[0]
        first_value = "" if raw_first is None else str(raw_first).strip()
        if first_value == _TOTAL_LABEL:
            continue
        for col_idx, symbol in symbol_columns:
            value = row[col_idx] if col_idx < len(row) else None
            count = _parse_cell_int(value)
            if count <= 0:
                continue
            aggregated[symbol] = aggregated.get(symbol, 0) + count
    return aggregated


def drawing_symbol_totals(
    path: str | None = None,
    category: str | None = None,
    drawings: list[str] | None = None,
) -> dict[str, dict[str, int]]:
    """{도면명: {부호: 도면 전체 합계}} 형태로 정답지를 파싱해 반환한다.

    Args:
        path: xlsx 경로 (None 이면 ANSWER_KEY_PATH).
        category: `"기둥"` / `"보"` / None.
            "기둥"·"보" 면 해당 접미사 시트만 선택, None 이면 같은 도면의
            기둥·보 시트를 dict merge.
        drawings: 도면명 화이트리스트 (예: ["도면1", "도면2", "도면4"]).
            None 이면 전체.

    합계 행(1열 == "합계")은 SUM 수식이라 캐시값을 신뢰하지 않고 건너뛴다.
    데이터 행에서 부호별 셀을 직접 누적 합산하고, 한 번이라도 0보다 큰
    값으로 등장한 부호만 결과 dict에 포함한다.
    """
    workbook = openpyxl.load_workbook(path or ANSWER_KEY_PATH, data_only=True)
    drawings_set = set(drawings) if drawings is not None else None
    totals: dict[str, dict[str, int]] = {}

    for sheet_name in workbook.sheetnames:
        drawing_name, sheet_category = _drawing_name_from_sheet(sheet_name)
        if category is not None and sheet_category != category:
            continue
        if drawings_set is not None and drawing_name not in drawings_set:
            continue

        aggregated = _parse_sheet(workbook[sheet_name])
        if not aggregated:
            continue

        bucket = totals.setdefault(drawing_name, {})
        for symbol, count in aggregated.items():
            bucket[symbol] = bucket.get(symbol, 0) + count

    return totals


def within_tolerance(
    predicted: int,
    expected: int,
    rel_tol: float = 0.05,
    small_count: int = 5,
) -> bool:
    """예측 개수가 정답 허용 오차 안에 있는지 판정.

    expected 가 small_count(기본 5) 이하면 ±1 까지 허용한다(작은 수에서
    5%는 너무 빡빡함). 그 외에는 상대오차 rel_tol(기본 5%) 이하면 통과.
    """
    diff = abs(predicted - expected)
    if expected <= small_count:
        return diff <= 1
    return diff <= expected * rel_tol
