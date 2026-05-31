"""기둥 규격 정답지 파서 — 라운드 길이-4.

`reference_materials/도면_길이_정답지.xlsx` 의 `도면N-기둥-길이` 시트
`비고` 컬럼에서 부호별 규격을 파싱한다.

비고 셀 형식
    "{부호} 규격: {규격원문} [(메모)]"
    예) "MC1 규격: H-588x300x12/20"
        "C1 규격: 600x407x20x35 (현장제작)"
        "SC1 규격: H 350x175x7/11"

원본 보존
    `spec_raw` 는 일람표에 적힌 그대로(공백·H 접두사 유무 포함).
    `spec_normalized` 는 단위중량 룩업용 — `normalize_spec()` 규칙으로 변환.

LLM·외부 호출 없는 결정론적 파서.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Optional

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(_HERE))
ANSWER_KEY_PATH = os.path.join(
    PROJECT_ROOT, "reference_materials", "도면_길이_정답지.xlsx"
)

_COLUMN_SUFFIX = "-기둥-길이"

_H_DRAWING = "도면"
_H_SECTION = "동·구역"
_H_SYMBOL = "부호"
_H_REMARK = "비고"
_H_LENGTH = "길이(mm)"
_H_INSTANCE = "인스턴스ID"

_REMARK_PATTERN = re.compile(
    r"^\s*(?P<symbol>[A-Za-z]+\d+)\s*규격\s*[:：]\s*"
    r"(?P<spec>[^()]+?)"
    r"(?:\s*\(\s*(?P<note>[^)]+?)\s*\))?\s*$"
)


@dataclass(frozen=True)
class SpecAnswer:
    """정답지의 (도면, 동, 부호) 단위 규격."""
    drawing: str
    section: Optional[str]
    symbol: str
    spec_raw: str
    spec_normalized: str
    spec_note: Optional[str]


@dataclass(frozen=True)
class InstanceAnswer:
    """정답지의 (도면, 동, 부호) 단위 인스턴스 집계.

    `count` 는 같은 키로 묶이는 행 개수, `length_mm` 은 첫 유효 길이(같은 키의
    인스턴스는 길이가 동일하다는 데이터 특성을 이용). 길이가 모든 행에서 None
    이면 `length_mm` 도 None — 길이 측정 불가 상태로 표기.
    """
    drawing: str
    section: Optional[str]
    symbol: str
    count: int
    length_mm: Optional[float]


def normalize_spec(spec_raw: str) -> str:
    """일람표 원문을 단위중량 룩업 키 형식으로 정규화.

    규칙
        - 공백 제거 ("H 350x175x7/11" → "H350x175x7/11")
        - H 접두사 자동 부여 ("600x407..." → "H-600x407...")
        - "H" 다음 하이픈 보장 ("H350..." → "H-350...")
        - 대문자 X·×  → 소문자 x 로 통일
        - 두께 구분자 슬래시 `/` → `x` 로 통일 (단, 첫 H- 하이픈은 보존)
          정답지 원문이 같은 단면을 슬래시·x 로 섞어 표기하는 경우가 있어
          (예: 도면1 2동 MC3 = "H-250x250x9x14" vs 일람표 DXF 의 "9/14")
          어느 표기든 동일 키로 룩업되도록 강제 통일.
    """
    if not spec_raw:
        return ""
    text = re.sub(r"\s+", "", spec_raw)
    text = text.replace("X", "x").replace("×", "x")
    if not text:
        return ""
    if text[0] in ("H", "h"):
        body = text[1:]
        if body.startswith("-"):
            body = body[1:]
        prefix = "H-"
    else:
        body = text
        prefix = "H-"
    # 본문(body) 내부의 슬래시만 x 로 치환. prefix 의 하이픈은 보존.
    body = body.replace("/", "x")
    return prefix + body


def _is_data_row(first_cell: object) -> bool:
    if first_cell is None:
        return False
    s = str(first_cell).strip()
    return s.startswith("도면") and len(s) > 2


def _build_header_index(header_row: tuple) -> dict[str, int]:
    return {
        str(c).strip(): i
        for i, c in enumerate(header_row)
        if c is not None and str(c).strip()
    }


def _parse_remark(remark: object) -> Optional[tuple[str, str, Optional[str]]]:
    """비고 셀 → (부호, 규격원문, 메모). 형식 불일치면 None."""
    if remark is None:
        return None
    text = str(remark).strip()
    if not text:
        return None
    match = _REMARK_PATTERN.match(text)
    if not match:
        return None
    return (
        match.group("symbol").strip(),
        match.group("spec").strip(),
        match.group("note").strip() if match.group("note") else None,
    )


def load_ground_truth_spec(
    path: str | None = None,
    drawings: list[str] | None = None,
) -> dict[tuple[str, Optional[str], str], SpecAnswer]:
    """정답지 비고 → (도면, 동, 부호) 단위 SpecAnswer 매핑.

    같은 (도면, 동, 부호) 가 여러 인스턴스 행에 반복 등장하지만 규격은 동일
    하다는 데이터 특성을 이용해 첫 번째 유효 행을 채택. 행마다 불일치하면
    `ValueError` 로 정답지 자체를 점검하게 한다.
    """
    workbook = openpyxl.load_workbook(path or ANSWER_KEY_PATH, data_only=True)
    drawings_set = set(drawings) if drawings is not None else None

    result: dict[tuple[str, Optional[str], str], SpecAnswer] = {}

    for sheet_name in workbook.sheetnames:
        if not sheet_name.endswith(_COLUMN_SUFFIX):
            continue
        drawing = sheet_name[: -len(_COLUMN_SUFFIX)]
        if drawings_set is not None and drawing not in drawings_set:
            continue

        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = _build_header_index(rows[0])

        idx_drawing = header.get(_H_DRAWING)
        idx_section = header.get(_H_SECTION)
        idx_symbol = header.get(_H_SYMBOL)
        idx_remark = header.get(_H_REMARK)
        if any(v is None for v in (idx_drawing, idx_symbol, idx_remark)):
            continue

        for row in rows[1:]:
            if not row or idx_drawing >= len(row):
                continue
            if not _is_data_row(row[idx_drawing]):
                continue

            symbol_cell = row[idx_symbol] if idx_symbol < len(row) else None
            remark_cell = row[idx_remark] if idx_remark < len(row) else None
            section_cell = (
                row[idx_section]
                if idx_section is not None and idx_section < len(row)
                else None
            )

            parsed = _parse_remark(remark_cell)
            if parsed is None:
                continue
            remark_symbol, spec_raw, note = parsed

            sheet_symbol = str(symbol_cell).strip() if symbol_cell else ""
            effective_symbol = sheet_symbol or remark_symbol
            section = str(section_cell).strip() if section_cell else None
            key = (drawing, section, effective_symbol)

            answer = SpecAnswer(
                drawing=drawing,
                section=section,
                symbol=effective_symbol,
                spec_raw=spec_raw,
                spec_normalized=normalize_spec(spec_raw),
                spec_note=note,
            )
            existing = result.get(key)
            if existing is None:
                result[key] = answer
            else:
                if existing.spec_normalized != answer.spec_normalized:
                    raise ValueError(
                        f"정답지 {sheet_name!r} 비고 불일치: {key} "
                        f"기존 {existing.spec_normalized!r} vs 신규 "
                        f"{answer.spec_normalized!r}"
                    )

    return result


def _coerce_length(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def load_section_instances(
    path: str | None = None,
    drawings: list[str] | None = None,
) -> dict[tuple[str, Optional[str], str], InstanceAnswer]:
    """정답지 인스턴스 행을 (도면, 동, 부호) 키로 집계.

    `count` = 그 키에 해당하는 데이터 행 수.
    `length_mm` = 같은 키의 인스턴스 중 첫 유효 길이(전부 None 이면 None).

    1단계 결과는 도면 단위 합계만 알려주므로, 라운드 길이-4 의 (도면,동,부호)
    단위 총중량 산출에는 별도의 per-section 카운트가 필요하다. 정답지에서
    인스턴스 단위로 직접 집계해 임시 카운트 소스로 쓴다.
    """
    workbook = openpyxl.load_workbook(path or ANSWER_KEY_PATH, data_only=True)
    drawings_set = set(drawings) if drawings is not None else None

    bucket: dict[tuple[str, Optional[str], str], list[Optional[float]]] = {}

    for sheet_name in workbook.sheetnames:
        if not sheet_name.endswith(_COLUMN_SUFFIX):
            continue
        drawing = sheet_name[: -len(_COLUMN_SUFFIX)]
        if drawings_set is not None and drawing not in drawings_set:
            continue

        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = _build_header_index(rows[0])

        idx_drawing = header.get(_H_DRAWING)
        idx_section = header.get(_H_SECTION)
        idx_symbol = header.get(_H_SYMBOL)
        idx_length = header.get(_H_LENGTH)
        if any(v is None for v in (idx_drawing, idx_symbol)):
            continue

        for row in rows[1:]:
            if not row or idx_drawing >= len(row):
                continue
            if not _is_data_row(row[idx_drawing]):
                continue
            symbol_cell = row[idx_symbol] if idx_symbol < len(row) else None
            section_cell = (
                row[idx_section]
                if idx_section is not None and idx_section < len(row)
                else None
            )
            length_cell = (
                row[idx_length]
                if idx_length is not None and idx_length < len(row)
                else None
            )

            if symbol_cell is None:
                continue
            symbol = str(symbol_cell).strip()
            if not symbol:
                continue
            section = str(section_cell).strip() if section_cell else None
            key = (drawing, section, symbol)
            bucket.setdefault(key, []).append(_coerce_length(length_cell))

    result: dict[tuple[str, Optional[str], str], InstanceAnswer] = {}
    for key, lengths in bucket.items():
        valid_lengths = [length for length in lengths if length is not None]
        first_length = valid_lengths[0] if valid_lengths else None
        result[key] = InstanceAnswer(
            drawing=key[0],
            section=key[1],
            symbol=key[2],
            count=len(lengths),
            length_mm=first_length,
        )
    return result
