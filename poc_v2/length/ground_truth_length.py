"""기둥 길이 정답지(xlsx) 로더 — 라운드 길이-1.

라운드 7 의 1단계 합계 정답지(`poc_v2/tests/ground_truth.py`)와 별개 워크북.
시트 구성:
    `도면N-기둥-길이`  — 도면별 기둥 인스턴스 단위 길이 정답
    `도면N-보-길이`    — 보 인스턴스 단위 길이 (본 라운드 범위 외)
    `도면라우팅`        — 측정 소스 도면 메타 (선택)
    `강재단위중량`      — 규격→단위중량 룩업 (선택)

시트 안 구조:
    1행 = 헤더 :
        `도면 | 동·구역 | 부호 | 인스턴스ID | 측정 소스 도면 | 도면 종류
        | 길이(mm) | 측정 근거 | 참조좌표 X | 참조좌표 Y | 신뢰도 | 비고`
    2행~ = 인스턴스 1개당 1행. `row[0]` 이 `도면N` 으로 시작.
    그 아래 = 시각적 구분선·부호별 요약 행. 데이터 행이 아니므로 무시.

셀 값 해석:
    `길이(mm)` 셀이 None / 비숫자 → "산출 불가" 또는 요약 헤더 → 제외.
    같은 (도면, 측정 소스, 부호) 묶음에 대해 모든 인스턴스 길이를 리스트로 모은다.

LLM·랜덤 요소 없는 결정론적 파서.
"""
from __future__ import annotations

import os
from typing import Optional

import openpyxl

# poc_v2/length/ground_truth_length.py → poc_v2/length → poc_v2 → 프로젝트 루트
_HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(_HERE))
ANSWER_KEY_PATH = os.path.join(
    PROJECT_ROOT, "reference_materials", "도면_길이_정답지.xlsx"
)

_COLUMN_SUFFIX = "-기둥-길이"
_ROUTING_SHEET = "도면라우팅"

# 정답지 1행에 기대하는 헤더 라벨
_H_DRAWING = "도면"
_H_SYMBOL = "부호"
_H_SOURCE = "측정 소스 도면"
_H_LENGTH = "길이(mm)"


def _build_header_index(header_row: tuple) -> dict[str, int]:
    """헤더 행에서 라벨→컬럼인덱스 매핑 생성."""
    return {
        str(cell).strip(): idx
        for idx, cell in enumerate(header_row)
        if cell is not None and str(cell).strip()
    }


def _is_data_row(first_cell: object) -> bool:
    """데이터 행 식별 — 첫 컬럼이 `도면N` 형태로 시작하는지."""
    if first_cell is None:
        return False
    text = str(first_cell).strip()
    return text.startswith("도면") and len(text) > 2


def _coerce_length(value: object) -> Optional[float]:
    """`길이(mm)` 셀 → float. 빈 셀·비숫자(예: 헤더 '길이(mm)')는 None."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):  # bool 은 int 서브타입 — 명시 제외
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def load_ground_truth_length(
    path: str | None = None,
    drawings: list[str] | None = None,
) -> dict[str, dict[str, dict[str, list[float]]]]:
    """기둥 길이 정답지를 파싱해 도면→측정소스→부호→길이리스트 형태로 반환.

    Args:
        path: xlsx 경로 (None 이면 ANSWER_KEY_PATH).
        drawings: 도면명 화이트리스트. None 이면 전체.

    Returns:
        `{도면명: {측정 소스 도면: {부호: [길이mm, ...]}}}`.

        길이가 None 인 행("산출 불가") 은 자동 제외된다 — 빈 (부호, 소스)
        쌍은 결과에 등장하지 않는다.
    """
    workbook = openpyxl.load_workbook(path or ANSWER_KEY_PATH, data_only=True)
    drawings_set = set(drawings) if drawings is not None else None

    result: dict[str, dict[str, dict[str, list[float]]]] = {}

    for sheet_name in workbook.sheetnames:
        if not sheet_name.endswith(_COLUMN_SUFFIX):
            continue
        drawing = sheet_name[: -len(_COLUMN_SUFFIX)]
        if drawings_set is not None and drawing not in drawings_set:
            continue

        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_index = _build_header_index(rows[0])
        try:
            idx_drawing = header_index[_H_DRAWING]
            idx_source = header_index[_H_SOURCE]
            idx_symbol = header_index[_H_SYMBOL]
            idx_length = header_index[_H_LENGTH]
        except KeyError as exc:
            raise ValueError(
                f"시트 {sheet_name!r} 에 필수 컬럼 누락: {exc!r}"
            ) from exc

        drawing_bucket = result.setdefault(drawing, {})
        for row in rows[1:]:
            if not row or idx_drawing >= len(row):
                continue
            if not _is_data_row(row[idx_drawing]):
                continue
            symbol = row[idx_symbol] if idx_symbol < len(row) else None
            source = row[idx_source] if idx_source < len(row) else None
            length = _coerce_length(
                row[idx_length] if idx_length < len(row) else None
            )

            if symbol is None or source is None:
                continue
            symbol = str(symbol).strip()
            source = str(source).strip()
            if not symbol or not source:
                continue
            if length is None:
                continue  # "산출 불가" 인스턴스 — 회귀 대상 아님

            source_bucket = drawing_bucket.setdefault(source, {})
            length_list = source_bucket.setdefault(symbol, [])
            length_list.append(length)

    return result


def load_routing_sheet(
    path: str | None = None,
) -> list[dict[str, object]]:
    """`도면라우팅` 시트를 dict 리스트로 반환.

    행 단위로 헤더→값 매핑. 본 라운드는 메타 정보용 — 실제 측정 라우팅은
    `config/length_routing.yaml` 을 사용한다.
    """
    workbook = openpyxl.load_workbook(path or ANSWER_KEY_PATH, data_only=True)
    if _ROUTING_SHEET not in workbook.sheetnames:
        return []

    worksheet = workbook[_ROUTING_SHEET]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    records: list[dict[str, object]] = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        record = {
            header[i]: row[i] if i < len(row) else None
            for i in range(len(header))
            if header[i]
        }
        records.append(record)
    return records
